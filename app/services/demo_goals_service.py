"""
Demo Goals Service

Fetches Sam's Club demo goals from productconnections.com via the Ninja Tables
AJAX API, filters by club number, and generates a formatted Excel file.
"""
import io
import re
import logging
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

DEMO_GOALS_PAGE_URL = 'https://productconnections.com/sams-club-demo-goals/'
AJAX_URL = 'https://productconnections.com/wp-admin/admin-ajax.php'
TABLE_ID = '5811'
NUM_CHUNKS = 5
_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
}


def _extract_nonce(html: str) -> str:
    """Extract the Ninja Tables public nonce from the page HTML."""
    match = re.search(r'ninja_table_public_nonce=([a-f0-9]+)', html)
    if not match:
        raise ValueError('Could not extract nonce from demo goals page')
    return match.group(1)


def _fetch_page_html(timeout: int = 30) -> str:
    """Fetch the demo goals page HTML to extract the nonce."""
    resp = requests.get(DEMO_GOALS_PAGE_URL, headers=_HEADERS, timeout=timeout, verify=False)
    resp.raise_for_status()
    return resp.text


def _fetch_all_rows(nonce: str, timeout: int = 30) -> list:
    """Fetch all data rows from the Ninja Tables AJAX API (all chunks)."""
    all_rows = []
    for chunk in range(NUM_CHUNKS):
        params = {
            'action': 'wp_ajax_ninja_tables_public_action',
            'table_id': TABLE_ID,
            'target_action': 'get-all-data',
            'default_sorting': 'manual_sort',
            'skip_rows': '0',
            'limit_rows': '0',
            'ninja_table_public_nonce': nonce,
            'chunk_number': str(chunk),
        }
        resp = requests.get(AJAX_URL, params=params, headers=_HEADERS, timeout=timeout, verify=False)
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list):
            all_rows.extend(data)
        else:
            logger.warning('Unexpected chunk %d response type: %s', chunk, type(data))
    return all_rows


def fetch_demo_goals(club_number: str, timeout: int = 30) -> list[dict]:
    """
    Fetch demo goals for a specific club number.

    Returns a list of dicts with keys:
        demoscheduleddate, demoid, itemnumber, itemdescription
    sorted by date ascending.
    """
    club_number = str(club_number).strip()

    html = _fetch_page_html(timeout=timeout)
    nonce = _extract_nonce(html)
    logger.info('Fetched nonce for demo goals: %s', nonce[:4] + '...')

    all_rows = _fetch_all_rows(nonce, timeout=timeout)
    logger.info('Fetched %d total rows from Ninja Tables', len(all_rows))

    # Filter by club number
    filtered = []
    for row in all_rows:
        val = row.get('value', {})
        if str(val.get('clubnumber', '')).strip() == club_number:
            filtered.append({
                'demoscheduleddate': val.get('demoscheduleddate', ''),
                'demoid': val.get('demoid', ''),
                'itemnumber': val.get('itemnumber', ''),
                'itemdescription': val.get('itemdescription', ''),
            })

    # Sort by Club Date ascending
    def parse_date(row):
        try:
            return datetime.strptime(row['demoscheduleddate'].strip(), '%m/%d/%Y')
        except (ValueError, AttributeError):
            return datetime.max

    filtered.sort(key=parse_date)
    logger.info('Found %d rows for club %s', len(filtered), club_number)
    return filtered


def generate_excel(rows: list[dict], club_number: str) -> bytes:
    """
    Generate a formatted .xlsx file from demo goals data.

    Layout:
        Row 1: Merged "Demo Goals" header (A1:D1)
        Row 2: Table headers (Club Date, Demo ID, Demo Item, Sales Target)
        Row 3+: Data rows sorted by date ascending
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Demo Goals'

    # Row 1: Merged header
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Demo Goals'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Column headers
    headers = ['Club Date', 'Demo ID', 'Demo Item', 'Sales Target']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    # Row 3+: Data
    for row_idx, row in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=row['demoscheduleddate'])
        ws.cell(row=row_idx, column=2, value=row['demoid'])
        ws.cell(row=row_idx, column=3, value=row['itemnumber'])
        ws.cell(row=row_idx, column=4, value=row['itemdescription'])

    last_row = max(2 + len(rows), 3)  # At least row 3 for valid table range

    # Create Excel Table (only if there's data)
    if rows:
        table_ref = f'A2:D{last_row}'
        tab = Table(displayName='DemoGoals', ref=table_ref)
        style = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Auto-fit column widths
    for col_idx in range(1, 5):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 4

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_demo_goals_excel(club_number: str, timeout: int = 30) -> tuple[bytes, str]:
    """
    Full pipeline: fetch, filter, and generate Excel.

    Returns (excel_bytes, filename).
    """
    rows = fetch_demo_goals(club_number, timeout=timeout)
    excel_bytes = generate_excel(rows, club_number)
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = f'Demo_Goals_{club_number}_{date_str}.xlsx'
    return excel_bytes, filename
